import copy
from itertools import product

from numpy.ma.extras import average

from old.environment import *
from generate_disruptions import *
from helper import *
import pandas as pd
import numpy as np
import gurobipy as gp
import os
import time
import pickle
import json
import random
import concurrent.futures
from scipy.stats import norm
from datetime import datetime, timedelta
from scipy.signal import savgol_filter
from gurobipy import GRB
from itertools import combinations
from concurrent.futures import ThreadPoolExecutor, as_completed
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error

# TODO:
# - Andere samenstelling features
# - Disruption costs as function of flight length?
# - individual probability of unavails?
#     - prob in aircraft state as well. # CHECK
#     - modify functions to work with probability of happening or not. #CHECK
#     - modify functions to calculate conflict probabilities.
# - Minder vluchten, OF cancellation costs hoger maken:
#       -Nu is cancellation bijna altijd de best optie,
#       -VFA model gaat daarnaast ook onnodig swappen dus mypic wint
#       - < 3.5 flights per aircraft.
# - Curfew later (+8 hours) CHECK


class VFA_ADP:
    def __init__(self, aircraft_data, flight_data, disruptions, recovery_start, recovery_end, agg_lvl, folder):
        self.folder = folder
        self.aircraft_data = aircraft_data
        self.flight_data = flight_data
        self.cancelled_flights = []

        self.aircraft_ids = [aircraft['ID'] for aircraft in self.aircraft_data]
        self.prone_aircraft = [self.aircraft_ids[0]]                         # Only first aircraft is prone to disruption
        self.recovery_start = recovery_start
        self.recovery_end = recovery_end

        self.interval = 60 # minutes
        self.intervals = pd.date_range(start=recovery_start, end=recovery_end, freq= str(self.interval)+'T')
        self.periods = {i: start for i, start in enumerate(self.intervals)}
        self.steps = [i for i in self.periods.keys()]
        self.period_length = pd.Timedelta(minutes=self.interval)
        self.total_recovery_time = (self.recovery_end - self.recovery_start).total_seconds() / 60  # Total recovery period in minutes

        self.potential_disruptions = []
        self.disruptions = disruptions
        self.d = 4                                              # Check "generate_disruptions.py" for disruption duration
        self.curfew = recovery_end + pd.Timedelta(hours=8)      # must be greater or equal then the max flight duration
        self.mu = len(self.steps[1:-2]) / 2 + 1                 # Center of the time horizon
        self.s = len(self.steps[1:-2]) / 3
        self.max_flights_per_aircraft_prone_ac = 6

        self.probabilities = norm.pdf(range(len(self.steps[1:-2])), self.mu, self.s)
        self.probabilities /= self.probabilities.sum()                                  # Normalize to sum to 1
        self.cumulative_probabilities = np.cumsum(self.probabilities)

        self.cancellation_cost = 300
        self.violation_costs = 150

        self.T = self.steps[-1]
        self.N = 1                          # Number of iterations per instance
        self.y = 1                          # Discount factor
        # self.a = 1 / self.N               # learning rate or stepsize, decaying
        self.α = 0.02                       # Learning rate or stepsize, fixed
        self.harmonic_a = self.N / 100     # Parameter for calculating harmonic stepsize
        self.ε = 0                          # Exploration probability
        self.ε_init = self.ε
        self.ε_end = 0.0

        self.harmonic_stepsize = True
        self.decaying_ε = True
        self.BFA = True
        self.pruning = False
        self.plot_vals = True
        self.plot_episode = False

        # States
        self.csv_file = '../_state_features_single_p.csv'
        self.states = dict()
        self.agg_states = dict()
        self.aggregation_level = agg_lvl

        # print(f'steps = {self.steps}')
        # print(f'periods = {self.periods}')

# INITIALIZATION FUNCTIONS:
    def initialize_state(self):
        t = 0
        current_time = self.periods[t]
        self.states = {}
        state_dict = dict()
        agg_dict = dict()

        state_dict['t'] = 0
        for aircraft in self.aircraft_data:
            aircraft_state = {}
            aircraft_id = aircraft['ID']
            ass_flights = self.ass_flights(aircraft_id, t)
            unavailibilty = []
            conflicts = []
            remaining_flights = len(ass_flights)

            aircraft_state = {'conflicts': conflicts, 'UA': unavailibilty, 'flights':ass_flights, 'n_remaining_flights': remaining_flights}
            state_dict[aircraft_id] = aircraft_state

        # initial_value = -self.cancellation_cost * self.num_conflicts(state_dict)
        # initial_value = -self.cancellation_cost * self.expected_num_conflicts(state_dict)
        initial_value = -self.initial_value(state_dict)
        # initial_value = 0

        # set value of the state to initial value and iteration to zero:
        state_dict['value'] = initial_value
        state_dict['iteration'] = 0

        agg_dict['count'] = 0
        agg_dict['value'] = initial_value
        # agg_dict['iteration'] = [0]

        state_key = self.create_hashable_state_key(state_dict)
        aggregate_state_key = self.G(state_dict, self.aggregation_level)

        self.states[state_key] = state_dict
        self.agg_states[aggregate_state_key] = agg_dict
        self.agg_states[aggregate_state_key]['count'] = 1

        return state_dict

    def ass_flights(self, aircraft_id, t):
        """Returns a list of copies flight dictionaries (with nr, adt, aat) assigned to an aircraft at a step, read from the data for intializing the states"""
        ass_flights = []
        flight_nrs = get_ac_dict(self.aircraft_data, aircraft_id)['AssignedFlights']
        for flight_nr in flight_nrs:
            flight = copy.deepcopy(get_flight_dict(flight_nr, self.flight_data))
            del flight["AssignedAircraft"]
            ass_flights.append(flight)
        return ass_flights

######### HELPER FUNCTIONS: #########
    def get_flight(self, flight_nr, current_state):
        for aircraft_id in self.aircraft_ids:
            aircraft_state = current_state[aircraft_id]
            for flight in aircraft_state['flights']:
                if flight['Flightnr'] == flight_nr:
                    return flight
        return 'None, Flight not found'

    #OLD FUNCTION - first level of aggregation
    def G(self, s, lvl):
        """
        Create a hashable aggregate key from a state dictionary that generalizes to unseen instances for testing purposes.

        The first level key includes:
        - n_remaining_flights per aircraft
        - rounded flight times to hours.

        The second level key includes:
        - n_remaining_flights per aircraft
        - n_remaining conflicts per aircraft

        Args:
            s (dict): A dictionary representing the state of the system.

        Returns:
            tuple: A hashable representation of the aggregate state.
        """
        aggregate_state_key = []

        # First level of aggregation
        if lvl == 1:
            # 1. Add the number of remaining flights for each aircraft
            for aircraft_id in self.aircraft_ids:
                aircraft_state = s[aircraft_id]
                # Create a list to store the aircraft attributes
                aircraft_key = []
                aircraft_key.append(aircraft_state['n_remaining_flights'])
                unavailability_times = []

                if aircraft_state['UA']:
                    unavailability_start = (aircraft_state['UA'][0][0] -  self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes
                    unavailability_end = (aircraft_state['UA'][0][1] -  self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes

                    unavailability_times.append(unavailability_start)
                    unavailability_times.append(unavailability_end)
                aircraft_key.append(tuple(unavailability_times))

                # 2. Round the flight times (departure and arrival) to the nearest hour
                rounded_flight_times = []

                # Sort the flights by departure time to ensure they are in chronological order
                sorted_flights = sorted(aircraft_state['flights'], key=lambda flight: flight['ADT'])

                for flight in sorted_flights:
                    # Round departure and arrival times to nearest hour
                    rounded_departure = flight['ADT'].replace(minute=0, second=0, microsecond=0) \
                                        + timedelta(hours=1 if flight['ADT'].minute >= 30 else 0)
                    rounded_arrival = flight['AAT'].replace(minute=0, second=0, microsecond=0) \
                                      + timedelta(hours=1 if flight['AAT'].minute >= 30 else 0)

                    # Relative arrtime and deptime to the recovery start time
                    dep_time = (rounded_departure - self.recovery_start).total_seconds() / 60 / self.total_recovery_time  # Convert to minutes
                    arr_time = (rounded_arrival - self.recovery_start).total_seconds() / 60 / self.total_recovery_time  # Convert to minutes

                    # Append the rounded times to the list
                    rounded_flight_times.append(dep_time)
                    rounded_flight_times.append(arr_time)

                # Add the number of remaining flights and the flight times (as a list, converted to tuple) for this aircraft
                aircraft_key.append(tuple(rounded_flight_times))
                aggregate_state_key.append(tuple(aircraft_key))

        # Second level of aggregation
        if lvl == 2:
            # 1. Add the number of remaining flights for each aircraft
            for aircraft_id in self.aircraft_ids:
                aircraft_state = s[aircraft_id]
                # Create a list to store the aircraft attributes
                aircraft_key = []
                n_conflicts = self.num_ac_conflicts(s, aircraft_id)

                aircraft_key.append(aircraft_state['n_remaining_flights'])
                aircraft_key.append(n_conflicts)

                aggregate_state_key.append(tuple(aircraft_key))

        # Convert the list to a tuple (to make it hashable)
        return tuple(aggregate_state_key)

    def initial_value(self, state):
        n_flights_dict = {}
        int1 = {}
        disruption = self.potential_disruptions[0][self.prone_aircraft[0]][0]      # Assume for disruption for one aircraft
        ua_start, ua_end = disruption[0], disruption[1]
        aircraft_overlaps = self.calculate_aircraft_overlaps(state)
        n_expected_conflicts = self.expected_num_conflicts(state)
        # print(f'n_expected_conflicts: {n_expected_conflicts}')

        for aircraft in self.aircraft_ids:
            aircraft_state = state[aircraft]
            n_flights_dict[aircraft] = len([f for f in aircraft_state['flights'] if f['AAT'] > ua_start])
            int1[aircraft] = aircraft_overlaps[aircraft] * n_flights_dict[aircraft]

        # print(f'min overlap: {min(int1.values())}')
        # print(f'min int1: {min(int1.values()) * n_expected_conflicts}')

        return min(int1.values()) * n_expected_conflicts

    # @profile
    def basis_features(self, state, x):
        '''time elapsed, times state is visisited, n_remaining flights, n_remaining conflicts, utilization, _disruption_occured, p'''
        features = {}
        # features['instance'] = self.folder
        features['t'] = state['t']
        features['count'] = state['iteration']
        # print(f't = {features['t']}')
        # print(f'count = {features['count']}')

        utilizations = {}
        total_conflicts = 0
        aircraft_overlaps = self.calculate_aircraft_overlaps(state)
        n_flights_dict = {}
        int1 = {}
        int2 = {}
        int3 = {}

        disruption_occured = self.check_disruption_(state['t'])
        E_n_conflicts = self.expected_num_conflicts(state)
        disruption = self.potential_disruptions[0][self.prone_aircraft[0]][0]      # Assume for disruption for one aircraft
        ua_start, ua_end = disruption[0], disruption[1]


        # 1. Add the number of remaining flights for each aircraft
        for aircraft_id in self.aircraft_ids:
            aircraft_state = state[aircraft_id]
            n_conflicts = self.num_ac_conflicts(state, aircraft_id)

            n_flights_dict[aircraft_id] =       len([f for f in aircraft_state['flights'] if f['AAT'] > ua_start])
            utilizations[aircraft_id] =         self.calculate_utilization(state, aircraft_id)
            int1[aircraft_id] =                 aircraft_overlaps[aircraft_id] *  n_flights_dict[aircraft_id]
            int2[aircraft_id] =                 utilizations[aircraft_id] *  n_flights_dict[aircraft_id]
            int3[aircraft_id] =                 aircraft_overlaps[aircraft_id] * utilizations[aircraft_id] *  n_flights_dict[aircraft_id]
            total_conflicts   +=                n_conflicts

            # if aircraft_id in self.prone_aircraft:
            #     flights = sorted(state[aircraft_id]['flights'], key=lambda f: f['ADT'])
            #
            #     # intialize probabilities at 0:
            #     for index in range(self.max_flights_per_aircraft_prone_ac):
            #         features[f'{aircraft_id}_F{index + 1}_prob'] = 0.0
            #
            #         # Now calculate probabilities for existing flights
            #         if index < len(flights):
            #             f = flights[index]
            #             features[f'{aircraft_id}_F{index + 1}_prob'] = self.individual_probability(f, aircraft_id, state)
            #             # a = features[f'{aircraft_id}_F{index + 1}_prob']
            #             # print(f' Prob {aircraft_id}, F{index+1} ={a}')
            #             # features[f'{aircraft_id}_F{index + 1}_prob'] = 1
            #
            #         # print(f'{aircraft_id}_F{index+1}_prob = {features[f'{aircraft_id}_F{index+1}_prob']}')

                # # Probabilities of different number of conflicts:
                # for n, prob in self.n_conflicts_probability_for_ac(flights, aircraft_id, state).items():
                #     if n > 5:
                #         continue
                #     features[f'{aircraft_id}_{n}C_prob'] = prob
                #     # print(f'{aircraft_id}_{n}C_prob = {prob}')

                # # Probabilities of different number of conflicts:
                # for n in range(1, 6):
                #     features[f'{aircraft_id}_{n}C_prob'] = 1
                #     # print(f'{aircraft_id}_{n}C_prob = {prob}')

        min_aircraft_id, min_overlap = min(aircraft_overlaps.items(), key=lambda x: x[1])

        min_util = min(utilizations.values())
        min_n_flights = min(n_flights_dict.values())
        min_int1 = min(int1.values())
        min_int2 = min(int2.values())
        min_int3 = min(int3.values())

        # print(f'{utilizations = }')
        # print(f'{n_flights_dict = }')
        # print(f'{aircraft_overlaps = }')
        # print(f'{int1 = }')
        # print(f'{int2 = }')
        # print(f'{int3 = }')

        features[f'min_prone_overlap'] =             min_overlap
        features[f'min_n_flights'] =                 min_n_flights
        features[f'min_util'] =                      min_util
        features[f'int1'] =                          min_int1                       # min(overlap        * n_flights)
        features[f'int2'] =                          min_int2                       # min(util           *  n_flights)
        features[f'int3'] =                          min_int3                       # min(overlap        * util         * n_flights)
        features[f'int4'] =                          E_n_conflicts * min_overlap    # min(E[n_conflicts] * overlap)

        features[f'E[n_conflicts]'] =                E_n_conflicts
        features[f'total_remaining_conflicts'] =     total_conflicts
        features[f'disruption_occured'] =            disruption_occured     # If a disruption occured
        features[f'recovered'] =                     1.0 if disruption_occured == 1 and total_conflicts == 0 else 0.0

        features['value'] = state['value']
        features['prev_action'] = x
        features['folder'] = self.folder

        return features

    # @profile
    def expected_num_conflicts(self, state):
        # return 0
        expected_disruptions = 0
        for aircraft in self.prone_aircraft:
            flights = sorted(state[aircraft]['flights'], key=lambda f: f['ADT'])

            # Calculate the individual disruption probabilities for each flight
            num_flights = len(flights)
            dep_times = [f['ADT'] for f in flights]
            result = {}

            probs = [self.individual_probability(f, aircraft, state) for f in flights if f['ADT'] >= self.periods[state['t']]]
            expected_disruptions += sum(probs)

        return expected_disruptions

    # @profile
    def individual_probability(self, flight, ac, state):
        t = state['t']
        p = 0.0
        for unavailability in self.potential_disruptions[t][ac]:
            start, end, prob, realises = unavailability
            if self.conflict_overlap(unavailability, flight):
                p = prob


        return float(p)

    def create_hashable_state_key(self, state_dict):
        """
        Create a hashable key from a state dictionary that generalizes to unseen instances for testing purposes.

        The key includes:
        - One entry for the current time step 't'
        - One entry for the remaining time left in the day
        - A sequence of departure and arrival times (in chronological order) for each aircraft.

        Args:
            state_dict (dict): A dictionary representing the state of the system.

        Returns:
            tuple: A hashable representation of the state.
        """
        state_key = []

        # 1. Add the current step 't' to the state key
        state_key.append(state_dict['t'])

        # 2. Calculate and add the remaining time left in the day
        current_time = self.periods[state_dict['t']]
        remaining_time = (self.recovery_end - current_time).total_seconds() / 60 / self.total_recovery_time # Time left in minutes
        state_key.append(remaining_time)

        # 3. For each aircraft, add the sequence of departure and arrival times
        for aircraft_id in self.aircraft_ids:
            aircraft_state = state_dict[aircraft_id]
            # Create a list to store the aircraft attributes
            aircraft_key = []
            aircraft_key.append(aircraft_state['n_remaining_flights'])
            unavailability_times = []

            if aircraft_state['UA']:
                unavailability_start = (aircraft_state['UA'][0][0] -  self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes
                unavailability_end = (aircraft_state['UA'][0][1] -  self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes

                unavailability_times.append(unavailability_start)
                unavailability_times.append(unavailability_end)
            aircraft_key.append(tuple(unavailability_times))

            flight_times = []
            # Sort the flights by departure time to ensure they are in chronological order
            sorted_flights = sorted(aircraft_state['flights'], key=lambda flight: flight['ADT'])

            for flight in sorted_flights:
                # Append both the departure and arrival times relative to the recovery start time
                dep_time = (flight['ADT'] - self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes
                arr_time = (flight['AAT'] - self.recovery_start).total_seconds() / 60 / self.total_recovery_time # Convert to minutes

                # Add both departure and arrival times to the sequence
                flight_times.append(dep_time)
                flight_times.append(arr_time)

            # Append the flight sequence to the state key for this aircraft
            aircraft_key.append(tuple(flight_times))
            state_key.append(tuple(aircraft_key))

        # Convert the list of state elements to a tuple (making it hashable)
        return tuple(state_key)

    def num_ac_conflicts(self, state, id):
        """Counts the total number of future conflicts for one aircraft id state"""
        n = 0
        current_time = self.periods[state['t']]
        for conflict in state[id]['conflicts']:
            if conflict['ADT'] >= current_time:
                n += 1
        return n

    def num_conflicts(self, state):
        """Counts the total number of future conflicts for the whole state"""
        n = 0
        current_time = self.periods[state['t']]
        for id in self.aircraft_ids:
            conflicts = state[id]['conflicts']
            unavailabilities = state[id]['UA']
            for conflict in conflicts:
                for ua in unavailabilities:
                    if ua[0] <= conflict['ADT'] >= current_time:
                        n += 1
        return n

    def num_remaing_flights(self, next_state, aircraft_id, next_step):
        n = 0
        next_time = self.periods[next_step]
        for flight in next_state[aircraft_id]['flights']:
            if flight['ADT'] >= next_time:
                n += 1
        return n

    def check_disruption_(self, t):
        """
        Check if a disruption occurred for a specific aircraft in a given iteration n.

        Parameters:
        disruptions (dict): Dictionary of disruptions by iteration.
        n (int): The current iteration to check.
        aircraft_id (str): The ID of the aircraft (e.g., '#1', '#2', '#3').

        Returns:
        int: 0 if a disruption occurred for the specified aircraft, 1 otherwise.
        """
        # Get the disruptions for the current iteration
        disruption_path = self.disruptions
        for aircraft in self.prone_aircraft:

            # Check each time step for any disruption event for the specified aircraft
            for time_step, disruption in disruption_path.items():
                if time_step < t:

                    # Check if there is any disruption event for the specified aircraft
                    if disruption[aircraft] != [] and disruption[aircraft][0][3]:
                        return 1         # Disruption found for this aircraft

        return 0  # No disruption found occured this aircraft

########################### LOGIC ###########################
    def unavailable_for_flight(self, flight, aircraft_id, t):
        """ Checks if the aircraft is unavailable due to disruption at Departure time of flight"""
        for disruption in self.disruptions[t][aircraft_id]:
            return (disruption[0] <= flight['ADT'] <= disruption[1] or
                    flight['ADT'] <= disruption[0] <= flight['AAT'])

    def conflict_at_step(self, state, aircraft_id, t):
        aircraft_state = state[aircraft_id]
        conflicts = []
        for flight in aircraft_state['flights']:
            for ua in aircraft_state['UA']:
                if self.disrupted(ua, flight):
                    conflicts.append(flight)
        return conflicts

    def overlaps(self, flight_to_swap, flight):
        """Checks if two flight times overlap"""
        return (flight_to_swap['ADT'] <= flight['ADT'] <= flight_to_swap['AAT'] or
                flight_to_swap['ADT'] <= flight['AAT'] <= flight_to_swap['AAT'] or
                (flight_to_swap['ADT'] <= flight['ADT'] and flight_to_swap['AAT'] >= flight['AAT']) or
                (flight_to_swap['ADT'] >= flight['ADT'] and flight_to_swap['AAT'] <= flight['AAT'])
                )

    def disruption_overlap(self, unavailability, flight):
        """Checks if flight overlaps with unavailability"""
        start = unavailability[0]
        end = unavailability[1]

        return (start <= flight['ADT'] <= end or
                start <= flight['AAT'] <= end or
                (start <= flight['ADT'] and end >= flight['AAT']) or
                (start >= flight['ADT'] and end <= flight['AAT'])
                )

    def conflict_overlap(self, unavailability, flight):
        """Checks if flight overlaps with unavailability"""
        start = unavailability[0]
        end = unavailability[1]

        return start <= flight['ADT'] < end

    def calculate_overlap_duration(self, unavailability, flight):
        """
        Calculates the total overlap duration between a flight and a disruption.

        Args:
            unavailability (tuple): A tuple with start and end times of the disruption (start, end).
            flight (dict): A dictionary with flight start ('ADT') and end ('AAT') times.

        Returns:
            pd.Timedelta: The duration of the overlap between the flight and disruption.
                          Returns a Timedelta of 0 if there is no overlap.
        """
        # Define start and end times for the unavailability (disruption)
        disruption_start = unavailability[0]
        disruption_end = unavailability[1]

        # Define start and end times for the flight
        flight_start = flight['ADT']
        flight_end = flight['AAT']

        # Calculate the overlap period
        overlap_start = max(disruption_start, flight_start)
        overlap_end = min(disruption_end, flight_end)

        # If overlap_start is earlier than overlap_end, there is an overlap
        if overlap_start < overlap_end:
            return overlap_end - overlap_start
        else:
            # No overlap
            return pd.Timedelta(0)

    def disrupted(self, unavailability, flight):
        """Checks if flight cannot depart due to unavailability,
            if flight already departed, the flight is not conflicted"""
        start = unavailability[0]
        end = unavailability[1]

        return (start <= flight['ADT'] < end or
                (start <= flight['ADT'] and end > flight['AAT']))

    def X_ta(self, current_state):
        '''
        :param current_state:
        :return: set of possible actions
        '''
        t = current_state['t']
        current_time = self.periods[current_state['t']]

        # Initialize actions with a default 'none' action
        actions = np.array([('none', 'none', 'none')], dtype=object)

        # Iterate over all aircraft
        for aircraft_id in self.aircraft_ids:
            aircraft_state = current_state[aircraft_id]

            # Iterate over all flights of the current aircraft
            for flight in aircraft_state['flights']:
                flight_nr = flight['Flightnr']
                flight_to_swap = self.get_flight(flight_nr, current_state)  # flight dict

                # Only consider flights that have not yet departed
                if flight_to_swap['ADT'] < current_time:
                    continue

                # Create cancellation action:
                new_action = np.array([('cancel', flight_nr, 'none')], dtype=object)
                actions = np.vstack([actions, new_action])

                # Consider swapping this flight to every other aircraft
                for other_aircraft_id in self.aircraft_ids:
                    if any(self.disrupted(unavail, flight) for unavail in current_state[other_aircraft_id]['UA']) and other_aircraft_id != aircraft_id:
                        continue
                    if self.calculate_utilization(current_state, other_aircraft_id) > 0.8 and self.pruning:
                        continue

                    # Create the swap action and append it to the actions array
                    new_action = np.array([('swap', flight_nr, other_aircraft_id)], dtype=object)
                    actions = np.vstack([actions, new_action])

        return actions

    ####################### REWARD LOGIC ########################
    def check_overlapping_assignments(self, next_state, aircraft_id, changed_flight):
        overlapping_flights = []

        # Check if the swapped flight overlaps with other flights assigned to the aircraft
        for flight in next_state[aircraft_id]['flights']:
            if flight != changed_flight and self.overlaps(changed_flight, flight):
                overlapping_flights.append(flight)

        # Sort the list of overlapping flights by ADT (Actual Departure Time)
        overlapping_flights = sorted(overlapping_flights, key=lambda flight: flight['ADT'])

        return overlapping_flights  # Return sorted overlapping flights

    def calculate_utilization(self, current_state, aircraft_id):
        """
        Calculates the utilization of an aircraft as the total occupied time (flights + disruptions)
        divided by the total remaining time until curfew.

        Args:
            current_time (pd.Timestamp): The current time.
            curfew_time (pd.Timestamp): The curfew time by which the aircraft must be free.
            flights (list): A list of dictionaries for flights, each with 'ADT' (departure) and 'AAT' (arrival).
            disruptions (list): A list of tuples, each with (start_time, end_time) for disruptions.

        Returns:
            float: The utilization ratio, capped at 1.0.
        """

        t = current_state['t']
        current_time = self.periods[current_state['t']]
        aircraft_state = current_state[aircraft_id]
        flights = aircraft_state['flights']
        disruptions = self.disruptions[t][aircraft_id]
        curfew_time = self.curfew
        # Calculate total remaining time until curfew
        total_remaining_time = max(pd.Timedelta(0), curfew_time - current_time)

        if total_remaining_time == pd.Timedelta(0):
            # If there's no remaining time until curfew, utilization is 0 by definition
            return 0.0

        # Collect all occupied periods (flights and disruptions) in a list of (start, end) tuples
        occupied_periods = []

        # Add flights as occupied periods
        for flight in flights:
            departure_time = max(flight['ADT'], current_time)  # Flight cannot start before current time
            arrival_time = min(flight['AAT'], curfew_time)  # Flight cannot end after curfew time
            if departure_time < arrival_time:
                occupied_periods.append((departure_time, arrival_time))

        # Add disruptions as occupied periods
        for disruption_start, disruption_end, p, realises in disruptions:
            if not realises:
                continue

            start_time = max(disruption_start, current_time)  # Disruption cannot start before current time
            end_time = min(disruption_end, curfew_time)  # Disruption cannot end after curfew time
            if start_time < end_time:
                occupied_periods.append((start_time, end_time))

        # Merge overlapping or contiguous periods to avoid double-counting
        occupied_periods.sort()  # Sort periods by start time
        merged_periods = []
        for start, end in occupied_periods:
            if merged_periods and merged_periods[-1][1] >= start:
                # If the current period overlaps or is contiguous with the last, merge them
                merged_periods[-1] = (merged_periods[-1][0], max(merged_periods[-1][1], end))
            else:
                # Otherwise, add the period as a new merged period
                merged_periods.append((start, end))

        # Calculate total occupied time by summing durations of merged periods
        total_occupied_time = sum((end - start for start, end in merged_periods), pd.Timedelta(0))

        # Calculate utilization as the ratio of occupied time to remaining time until curfew
        utilization = total_occupied_time / total_remaining_time

        return float(utilization)

    def calculate_aircraft_overlaps(self, state):
        t = state['t']
        aircraft_overlaps = {}
        for aircraft in self.prone_aircraft:
            disrupted_flights = [f for f in state[aircraft]['flights'] if any(self.disrupted(u, f) for u in self.potential_disruptions[t][aircraft] if u[2] != 0)]

            # Check is flight can be delayed by checking overlap length with end of disruption (overlap with itself)
            ua_overlaps = [max((ua[1] - f['ADT']).total_seconds() / 60  ,0.0) for f in disrupted_flights for ua in self.potential_disruptions[t][aircraft]]
            ua_overlap = sum(ua_overlaps)
            aircraft_overlaps[aircraft] = ua_overlap

            for other_aircraft in self.aircraft_ids:
                if other_aircraft != aircraft:
                    aircraft_overlaps[other_aircraft] = 0.0
                    flights_overlap = 0.0

                    for f in disrupted_flights:
                        disrupted_flight_start = f['ADT']
                        disrupted_flight_end = f['AAT']

                        for other_flight in state[other_aircraft]['flights']:
                            other_start = other_flight['ADT']
                            other_end = other_flight['AAT']

                            overlap_start = max(disrupted_flight_start, other_start)
                            overlap_end = min(disrupted_flight_end, other_end)

                            if overlap_start < overlap_end:  # If there's a valid overlap
                                overlap_minutes = (overlap_end - overlap_start).total_seconds() / 60
                                flights_overlap += overlap_minutes

                    aircraft_overlaps[other_aircraft] += flights_overlap  # Add flight overlap to the total

        return aircraft_overlaps

    def delay_swapped_flight(self, next_state, aircraft_id, changed_flight, overlapping_flights, apply):
        """
        Delay the swapped flight and potentially other flights to prevent overlap.

        Args:
            next_state (dict): The state of the system after swapping.
            aircraft_id (str): The ID of the aircraft.
            changed_flight (dict): The swapped flight.
            overlapping_flights (list): A list of flights that overlap with the swapped flight.

        Returns:
            dict or None: The updated next state if conflicts are resolved, otherwise None.
        """
        temp_next_state = copy.deepcopy(next_state)
        step = next_state['t']
        if not overlapping_flights:
            return None

        # Process each overlapping flight
        while overlapping_flights:
            if apply:
                print(f'\n##################################')
                print(f'Changed Flight = {changed_flight['Flightnr']} is assigned to {aircraft_id}')
                print(f'{[flight["Flightnr"] for flight in overlapping_flights]}')


            for overlapping_flight in overlapping_flights:
                # Case 1: Swapped/changed flight departs later, delay swapped/changed flight
                if overlapping_flight['ADT'] < changed_flight['ADT']:
                    new_start_time = overlapping_flight['AAT'] + pd.Timedelta(minutes=10)
                    flight_duration = changed_flight['AAT'] - changed_flight['ADT']
                    new_end_time = new_start_time + flight_duration

                    if apply:
                        print(f'Delayed flight {changed_flight["Flightnr"]} with {new_start_time - changed_flight["ADT"]}')


                    # Apply the delay to the swapped/changed flight
                    changed_flight['ADT'] = new_start_time
                    changed_flight['AAT'] = new_end_time
                    delayed_flight = changed_flight

                    # IMPORTANT: Update the flight in the temp_next_state
                    for i, flight in enumerate(temp_next_state[aircraft_id]['flights']):
                        if flight['Flightnr'] == delayed_flight['Flightnr']:
                            temp_next_state[aircraft_id]['flights'][i] = delayed_flight

                # Case 2: Overlapping flight departs later, delay the overlapping flight
                else:
                    new_start_time = changed_flight['AAT'] + pd.Timedelta(minutes=10)
                    flight_duration = overlapping_flight['AAT'] - overlapping_flight['ADT']
                    new_end_time = new_start_time + flight_duration
                    if apply:
                        print(f'Delayed flight {overlapping_flight["Flightnr"]} with {new_start_time - overlapping_flight["ADT"]}')

                    # Apply the delay to the overlapping flight
                    overlapping_flight['ADT'] = new_start_time
                    overlapping_flight['AAT'] = new_end_time
                    delayed_flight = overlapping_flight

                    # IMPORTANT: Update the flight in the temp_next_state
                    for i, flight in enumerate(temp_next_state[aircraft_id]['flights']):
                        if flight['Flightnr'] == delayed_flight['Flightnr']:
                            temp_next_state[aircraft_id]['flights'][i] = delayed_flight

            # Re-check for new conflicts after each delay
            if apply:
                print(f'Check overlapping assignment for newly changed flight: {delayed_flight}')
            overlapping_flights = self.check_overlapping_assignments(temp_next_state, aircraft_id, delayed_flight)
            if apply:
                print(f'Overlapping flights:')
                print(f'{[flight for flight in overlapping_flights]}')
            changed_flight = delayed_flight

            # If new conflicts emerge, continue the loop, otherwise break
            if not overlapping_flights or overlapping_flights == []:
                break

        # next_state = temp_next_state
        # If no further conflicts, return the updated state and total delay
        return temp_next_state

    def delay_disrupted_flight(self, next_state, aircraft_id, disrupted_flight, unavailability_periods, apply):
        """
        Delay the disrupted flight due to aircraft unavailability and propagate delays for subsequent flights.

        Args:
            next_state (dict): The state of the system after the disruption.
            aircraft_id (str): The ID of the aircraft.
            disrupted_flight (dict): The disrupted flight due to aircraft unavailability.
            unavailability_periods (list): A list of (start_time, end_time) tuples for the aircraft's unavailability periods.
            apply (bool): If True, print debug information.

        Returns:
            dict: The updated next state with the delayed flight and any subsequent delays.
        """
        temp_next_state = copy.deepcopy(next_state)
        step = next_state['t']

        if not unavailability_periods:
            return None

        # Sort unavailability periods to ensure processing them in order
        unavailability_periods.sort(key=lambda x: x[1])  # Sort by end_time

        # Process each unavailability period
        for unavailability in unavailability_periods:
            unavailability_start, unavailability_end, p, realises = unavailability
            if not realises:
                continue

            # Check if the flight's ADT is within the unavailability period
            if self.disruption_overlap(unavailability, disrupted_flight):
                # Delay the disrupted flight to after the unavailability ends
                new_start_time = unavailability_end + pd.Timedelta(minutes=10)  # Add a buffer of 10 minutes
                flight_duration = disrupted_flight['AAT'] - disrupted_flight['ADT']
                new_end_time = new_start_time + flight_duration

                if apply:
                    print(f'Delayed flight {disrupted_flight["Flightnr"]} to start after unavailability ends at {new_start_time}')
                    print(f'UA: {unavailability_periods}')
                    print(f'Disrupted flight: {disrupted_flight}')

                # Apply the delay to the disrupted flight
                disrupted_flight['ADT'] = new_start_time
                disrupted_flight['AAT'] = new_end_time

                # IMPORTANT: Update the flight in the temp_next_state
                for i, flight in enumerate(temp_next_state[aircraft_id]['flights']):
                    if flight['Flightnr'] == disrupted_flight['Flightnr']:
                        temp_next_state[aircraft_id]['flights'][i] = disrupted_flight

            # Check for subsequent flights and delay them as needed
            overlapping_flights = self.check_overlapping_assignments(temp_next_state, aircraft_id, disrupted_flight)
            if overlapping_flights != [] and self.delay_swapped_flight(temp_next_state, aircraft_id, disrupted_flight, overlapping_flights, apply=False) is not None:
                temp_next_state = self.delay_swapped_flight(temp_next_state, aircraft_id, disrupted_flight, overlapping_flights, apply=False)

        # Update the next state with all delayed flights
        # next_state = temp_next_state
        return temp_next_state

    def check_delays(self, current_state, next_state):
        total_delay = 0
        pre_flights = [flight for aircraft_id in self.aircraft_ids for flight in current_state[aircraft_id]['flights']]
        post_flights = [flight for aircraft_id in self.aircraft_ids for flight in next_state[aircraft_id]['flights']]

        pre_flights = sorted(pre_flights, key=lambda flight: flight['Flightnr'])
        post_flights = sorted(post_flights, key=lambda flight: flight['Flightnr'])

        for pre_flight, post_flight in zip(pre_flights, post_flights):
            # Calculate the delay in minutes (if any) by comparing ADT
            delay_minutes = (post_flight['ADT'] - pre_flight['ADT']).total_seconds() / 60.0
            if delay_minutes > 0:
                total_delay += delay_minutes

        return total_delay

    def check_canx(self, current_state, next_state, aircraft_id):
        current_t = current_state['t']
        if current_t == self.T:
            return 0

        next_t = next_state['t']
        current_time = self.periods[current_t]
        next_time = self.periods[next_t]

        canx = 0
        pre_conflicts = current_state[aircraft_id]['conflicts']
        post_conflicts = next_state[aircraft_id]['conflicts']

        # Check for conflicts that started in the current period and remain unresolved in the next period
        for conflict in pre_conflicts:
            # Conflicted flight departs before the next period and still exists in the next state
            if current_time <= conflict['ADT'] < next_time and conflict in post_conflicts:
                canx += 1

        return canx

    def check_curfew_violations(self, current_state, next_state, aircraft_id):
        current_t = current_state['t']
        violations = 0

        pre_violations = 0
        for flight in current_state[aircraft_id]['flights']:
            if flight['AAT'] > self.curfew:
                pre_violations += 1

        post_violations = 0
        for flight in next_state[aircraft_id]['flights']:
            if flight['AAT'] > self.curfew:
                post_violations += 1

        violations = max(0, (post_violations - pre_violations))

        return violations

    def compute_reward(self, pre_decision_state, post_decision_state, action):
        """
        Computes the reward for transitioning from the pre-decision state to the post-decision state.

        Args:
            pre_decision_state (dict): The state of the system before the action was taken.
            post_decision_state (dict): The state of the system after the action was taken.
            action (tuple): The action taken, which could be a swap or other types of actions.

        Returns:
            int: The computed reward.
        """
        reward = 0
        action_type, swapped_flight, old_aircraft_id = action

        canx = 0
        violations = 0
        for aircraft_id in self.aircraft_ids:
            # 1. Check how many flights did not get recoverd or violated curfews
            canx += self.check_canx(pre_decision_state, post_decision_state, aircraft_id)
            violations += self.check_curfew_violations(pre_decision_state, post_decision_state, aircraft_id)

        # Make sure that instead of do nothing, the model actively cancels the flight by choosing 'cancel' action
        # This way the model cancels the same flight by frees up space by doing so.
        reward -= canx * 10000
        reward -= violations * self.violation_costs

        # Penalties for performing a swap actions
        if action_type == 'swap':
            reward -= 10

            # Check if delays were necessary following the swaps:
            delay = self.check_delays(pre_decision_state, post_decision_state)
            reward -= delay

        if action_type == 'cancel':
            reward -= self.cancellation_cost

        # # add 10 rewards for time costs
        # reward -= 10

        return reward


    # @profile
    def simulate_action_to_state(self, S_t_dict, x, t, n):
        '''Does the same as apply action to state, run when checking actions. Run apply_action_to_state for actually appliying actions'''
        # Create a copy of the current state to modify
        S_tx_dict = copy.deepcopy(S_t_dict)
        current_step = S_t_dict['t']
        next_step = current_step + 1

        action_type, swapped_flight_nr, new_aircraft_id = x

        if action_type == 'cancel':
            # Find the flight in the current aircraft's state flight_nr
            old_aircraft_id = next((aircraft_id for aircraft_id, aircraft_state in S_tx_dict.items()
                                    if aircraft_id != 't' and
                                    aircraft_id != 'time_left' and
                                    any(flight['Flightnr'] == swapped_flight_nr for flight in aircraft_state['flights'])), None)
            old_aircraft_state = S_tx_dict[old_aircraft_id]
            flight_to_swap = next(flight for flight in old_aircraft_state['flights'] if flight['Flightnr'] == swapped_flight_nr)
            # Remove the flight from the old aircraft and assign it to the new aircraft
            S_tx_dict[old_aircraft_id]['flights'].remove(flight_to_swap)

        if action_type == 'swap':
            # 1. Swap the assignments of the aircraft for the flight
            old_aircraft_id = next((aircraft_id for aircraft_id, aircraft_state in S_tx_dict.items()
                                    if aircraft_id != 't' and
                                    aircraft_id != 'time_left' and
                                    any(flight['Flightnr'] == swapped_flight_nr for flight in aircraft_state['flights'])), None)
            old_aircraft_state = S_tx_dict[old_aircraft_id]

            # Find the flight in the current aircraft's stateflight_nr
            flight_to_swap = next(flight for flight in old_aircraft_state['flights'] if flight['Flightnr'] == swapped_flight_nr)

            # Remove the flight from the old aircraft and assign it to the new aircraft
            S_tx_dict[new_aircraft_id]['flights'].append(flight_to_swap)
            S_tx_dict[old_aircraft_id]['flights'].remove(flight_to_swap)

            # 2 Check for overlaps and delay flights if possible:
            swapped_flight = flight_to_swap
            overlapping_flights = self.check_overlapping_assignments(S_tx_dict, new_aircraft_id, swapped_flight)
            if overlapping_flights != [] and self.delay_swapped_flight(S_tx_dict, new_aircraft_id, swapped_flight, overlapping_flights, apply=False) is not None:
                S_tx_dict = self.delay_swapped_flight(S_tx_dict, new_aircraft_id, swapped_flight, overlapping_flights, apply=False)

            # 3 Check for unavailability and delay flights if possible
            unavailabilities = S_tx_dict[new_aircraft_id]['UA']
            if self.delay_disrupted_flight(S_tx_dict, new_aircraft_id, swapped_flight, unavailabilities, apply=False) is not None:
                S_tx_dict = self.delay_disrupted_flight(S_tx_dict, new_aircraft_id, swapped_flight, unavailabilities, apply=False)

        if action_type == 'none':
            # nothing happens to the assignments when doing nothing.
            pass

        for i, aircraft_id in enumerate(self.aircraft_ids):
            aircraft_state = S_tx_dict[aircraft_id]
            aircraft_state['conflicts'] = self.conflict_at_step(S_tx_dict, aircraft_id, next_step) if t != self.T else 0
            aircraft_state['n_remaining_flights'] = self.num_remaing_flights(S_tx_dict, aircraft_id, next_step)

        # Update time for next state:
        S_tx_dict['t'] = next_step

        S_tx = self.create_hashable_state_key(S_tx_dict)
        if S_tx in self.states:
            return self.states[S_tx]
        else:
            # if it is a newly expored state: calculated the intial value as function of t
            # S_tx_dict['value'] = -self.cancellation_cost * self.num_conflicts(S_tx_dict)
            S_tx_dict['value'] = -self.cancellation_cost * self.expected_num_conflicts(S_tx_dict)
            S_tx_dict['value'] = -self.initial_value(S_tx_dict)
            # S_tx_dict['value'] = [-self.cancellation_cost * (len(self.flight_data) / len(self.aircraft_ids))]
            S_tx_dict['iteration'] = 0
            return S_tx_dict

    def apply_action_to_state(self, S_t_dict, x, t, n):
        '''Does the same as apply action to state, run when checking actions. Run apply_action_to_state for actually appliying actions'''
        # Create a copy of the current state to modify
        S_tx_dict = copy.deepcopy(S_t_dict)
        current_step = S_t_dict['t']
        next_step = current_step + 1

        action_type, swapped_flight_nr, new_aircraft_id = x

        if action_type == 'cancel':
            # Find the flight in the current aircraft's state flight_nr
            old_aircraft_id = next((aircraft_id for aircraft_id, aircraft_state in S_tx_dict.items()
                                    if aircraft_id != 't' and
                                    aircraft_id != 'time_left' and
                                    any(flight['Flightnr'] == swapped_flight_nr for flight in aircraft_state['flights'])), None)
            old_aircraft_state = S_tx_dict[old_aircraft_id]
            flight_to_swap = next(flight for flight in old_aircraft_state['flights'] if flight['Flightnr'] == swapped_flight_nr)
            # Remove the flight from the old aircraft and assign it to the new aircraft
            S_tx_dict[old_aircraft_id]['flights'].remove(flight_to_swap)
            self.cancelled_flights.append((flight_to_swap, old_aircraft_id))

        if action_type == 'swap':
            # 1. Swap the assignments of the aircraft for the flight
            old_aircraft_id = next((aircraft_id for aircraft_id, aircraft_state in S_tx_dict.items()
                                    if aircraft_id != 't' and
                                    aircraft_id != 'time_left' and
                                    any(flight['Flightnr'] == swapped_flight_nr for flight in aircraft_state['flights'])), None)
            old_aircraft_state = S_tx_dict[old_aircraft_id]

            # Find the flight in the current aircraft's stateflight_nr
            flight_to_swap = next(flight for flight in old_aircraft_state['flights'] if flight['Flightnr'] == swapped_flight_nr)

            # Remove the flight from the old aircraft and assign it to the new aircraft
            S_tx_dict[new_aircraft_id]['flights'].append(flight_to_swap)
            S_tx_dict[old_aircraft_id]['flights'].remove(flight_to_swap)

            # 2 Check for overlaps and delay flights if possible:
            swapped_flight = flight_to_swap
            overlapping_flights = self.check_overlapping_assignments(S_tx_dict, new_aircraft_id, swapped_flight)
            if overlapping_flights != [] and self.delay_swapped_flight(S_tx_dict, new_aircraft_id, swapped_flight, overlapping_flights, apply=False) is not None:
                S_tx_dict = self.delay_swapped_flight(S_tx_dict, new_aircraft_id, swapped_flight, overlapping_flights, apply=False)

            # 3 Check for unavailability and delay flights if possible
            unavailabilities = S_tx_dict[new_aircraft_id]['UA']
            if self.delay_disrupted_flight(S_tx_dict, new_aircraft_id, swapped_flight, unavailabilities, apply=False) is not None:
                S_tx_dict = self.delay_disrupted_flight(S_tx_dict, new_aircraft_id, swapped_flight, unavailabilities, apply=False)

        if action_type == 'none':
            # nothing happens to the assignments when doing nothing.
            pass

        for i, aircraft_id in enumerate(self.aircraft_ids):
            aircraft_state = S_tx_dict[aircraft_id]
            aircraft_state['conflicts'] = self.conflict_at_step(S_tx_dict, aircraft_id, next_step) if t != self.T else 0
            aircraft_state['n_remaining_flights'] = self.num_remaing_flights(S_tx_dict, aircraft_id, next_step)

        # Update time for next state:
        S_tx_dict['t'] = next_step

        S_tx = self.create_hashable_state_key(S_tx_dict)
        if S_tx in self.states:
            return self.states[S_tx], S_tx
        else:
            # if it is a newly expored state: calculated the intial value as function of t
            # S_tx_dict['value'] = -self.cancellation_cost * self.num_conflicts(S_tx_dict)
            S_tx_dict['value'] = -self.cancellation_cost * self.expected_num_conflicts(S_tx_dict)
            S_tx_dict['value'] = -self.initial_value(S_tx_dict)
            # S_tx_dict['value'] = [-self.cancellation_cost * (len(self.flight_data) / len(self.aircraft_ids))]
            S_tx_dict['iteration'] = 0
            self.states[S_tx] = S_tx_dict
            return self.states[S_tx], S_tx

    def add_exogeneous_info(self, S_tx, n):
        S_tx_dict = self.states[S_tx]
        t = S_tx_dict['t']
        W_t_next = self.disruptions[t]
        S_t_next_dict = copy.deepcopy(S_tx_dict)

        for aircraft_id in self.aircraft_ids:
            aircraft_state = S_t_next_dict[aircraft_id]
            W_t_aircraft = W_t_next[aircraft_id]
            if W_t_aircraft != []:
                (start, end, p, realises) = W_t_aircraft[0]
                if not realises:
                    continue

            # Add Exogeneous information - new realizations of aircraft unavailabilities:
            S_t_next_dict[aircraft_id]['UA'] = W_t_aircraft

            # Update state attribute 'conflicts' for new information
            aircraft_state['conflicts'] = self.conflict_at_step(S_t_next_dict, aircraft_id, t) if t != self.T else aircraft_state['conflicts']
            aircraft_state['n_remaining_flights'] = self.num_remaing_flights(S_t_next_dict, aircraft_id, t)

        S_t_next = self.create_hashable_state_key(S_t_next_dict)
        # if not S_t_next in self.states:
        #     # if it is a newly expored state: calculated the intial value as function of t
        #     S_t_next_dict['value'] = [-self.cancellation_cost * self.num_conflicts(S_t_next_dict)]
        #     S_t_next_dict['value'] = [S_tx_dict['value'][-1]]
        #     S_t_next_dict['iteration'] = [n]

        return S_t_next_dict, S_t_next

################### SOLVE: ##################
    def solve_with_gurobi(self, S_t_dict, X_ta, t, n):
        # Initialize the Gurobi model
        model = gp.Model("ADP_Optimization")

        # Add decision variables for each action in X_ta
        action_vars = model.addVars(len(X_ta), vtype=GRB.BINARY, name="actions")

        # Objective function: maximize the expected value
        V_x = {}
        R_t = {}
        for i, x in enumerate(X_ta):
            S_tx_dict = self.simulate_action_to_state(S_t_dict, x, t, n)
            r_t = self.compute_reward(S_t_dict, S_tx_dict, x)
            v_downstream = S_tx_dict['value']
            V_x[i] = r_t + self.y * v_downstream
            R_t[tuple(x)] = r_t

        # Set the objective to maximize the sum of expected rewards for chosen actions
        model.setObjective(gp.quicksum(V_x[i] * action_vars[i] for i in range(len(X_ta))), GRB.MAXIMIZE)

        # Add any necessary constraints (e.g., constraints on actions)
        # For example: if only one action can be chosen at a time
        model.addConstr(gp.quicksum(action_vars[i] for i in range(len(X_ta))) == 1, "action_constraint")

        # Optimize the model
        model.optimize()

        # Extract the best action: find the action with a solution value of 1
        best_action_idx = None
        for i in range(len(X_ta)):
            if action_vars[i].x > 0.02:  # Gurobi returns solution values as floats (0.0 or 1.0)
                best_action_idx = i
                break

        if best_action_idx is not None:
            x_hat = X_ta[best_action_idx]
            v_hat = V_x[best_action_idx]
            return x_hat, v_hat, R_t
        else:
            raise ValueError("No valid action found")

    # @profile
    def train_with_vfa(self):
        objective_values = {}
        value_function_values = {}
        self.policy = {}
        agg_states_count = 1
        collected_features_list = []                                                     # List to collect state features as dictionaries for final iteration

        disruptions = load_disruptions("Disruptions_train") # get pre sampled disruptions from storage
        shuffled_disruption_paths = get_shuffled_disruption_paths(disruptions, self.N)
        self.disruptions = shuffled_disruption_paths[0]
        self.potential_disruptions = self.potential_and_realised_disruptions()
        print('self.disruptions:')
        print(self.disruptions)
        print()
        print('self.potential_disruptions:')
        print(self.potential_disruptions)

        self.initial_state = self.initialize_state()
        self.initial_state_key = self.create_hashable_state_key(self.initial_state)

        TIME = time.time()
        count = 0
        with open(self.csv_file, mode='a') as q:
            for n in range(1, int(self.N) + 1):
                self.disruptions = shuffled_disruption_paths[n-1]
                self.potential_disruptions = self.potential_and_realised_disruptions()
                next_state = self.states[self.initial_state_key]
                f = (self.N - n) / self.N                                               # Decaying ε
                ε = (self.ε_init - self.ε_end)*f + self.ε_end                           # Decaying ε
                initial_expected_value = next_state['value']
                # count += count_disruptions(self.disruptions)

                accumulated_rewards = []
                # state_features = self.basis_features(next_state)
                # if n == self.N:
                #     collected_features_list.extend([state_features])

                for t in self.steps[:-1]:
                    if n == self.N and self.plot_episode:
                        self.plot_schedule(next_state, n, self.folder)

                    V_x = {}
                    R_t = {}
                    V_downstream = {}

                    S_t_dict = next_state  # Pre-decsision state
                    S_t = self.create_hashable_state_key(S_t_dict)
                    S_tx_prev_dict = self.states[S_tx] if t > 0 else self.states[S_t]  # Carry the post-decision state forward from previous timestep (It gets updated at this step timestep)
                    S_tx_prev = S_tx if t > 0 else S_t

                    # print(f'\nS_tx[t-1]:')
                    # self.print_state(S_tx_prev_dict)

                    # S_tx_g_prev = self.G(S_tx_prev_dict, self.aggregation_level)
                    # S_tx_g_prev_dict = self.agg_states[S_tx_g_prev] if S_tx_g_prev in self.agg_states else {'count': 1, 'value': S_tx_prev_dict['value']}
                    # self.agg_states[S_tx_g_prev] = S_tx_g_prev_dict

    # CALCULATE STATES AND VALUES:
    #                 print(f'\nS_t:')
    #                 self.print_state(S_t_dict)
                    X_ta = self.X_ta(S_t_dict) # Action set np array
                    for x in X_ta:
                        S_tx_dict = self.simulate_action_to_state(S_t_dict, x, t, n)
                        r_t = self.compute_reward(S_t_dict, S_tx_dict, x)
                        v_downstream = S_tx_dict['value']
                        x = tuple(x)
                        V_downstream[x] = v_downstream
                        R_t[x] = r_t
                        V_x[x] = r_t + self.y * v_downstream

                    # if t == 0 and n == self.N:
                    #     for x in X_ta:
                    #         x = tuple(x)
                    #         print('--------------------------')
                    #         print(x)
                    #         print(R_t[x], V_downstream[x], V_x[x])

                    # Choose the best action
                    x_hat = max(V_x, key=V_x.get)
                    v_hat = V_x[x_hat]
                    # self.policy[S_t] = x_hat
                    best_immediate_reward = R_t[x_hat]
                    accumulated_rewards.append(best_immediate_reward)

                    # if n == self.N:
                    #     print(f' {t = }')
                    #     print(f'\n\n\tBEST ACTION AT {t = } >>>> {x_hat = }:')
                    #     print(f'\tBEST IMMEDIATE REWARD >>>> R_t {best_immediate_reward}:')
                    #     print(f'\tDOWNSTREAM REWARD     >>>> V(S_tx) = {V_downstream[x_hat]}:')
                    #     print(f'\tBEST ACTION VALUE     >>>> V(S_t){v_hat}')

    # UPDATE VALUES:
        # Update the value of the previous post-decision state
                    v_n_prev = self.states[S_tx_prev]['value'] # Value of last timestep post-decision state at previous iteration
                    if self.harmonic_stepsize:
                        N = self.states[S_tx_prev]['iteration']
                        α =  self.harmonic_a / (self.harmonic_a + n - 1)
                        v_n_new = (1 - α) * v_n_prev + α * v_hat  # Value of previous timestep post-decision state at this timestep
                        # print(f'{v_n_prev = }')
                        # print(f'v_n_new = (1 - {α})*{v_n_prev} + {α}*{v_hat} = {v_n_new}\n')
                    else:
                        v_n_new = (1 - self.α) * v_n_prev + self.α * v_hat # Value of previous timestep post-decision state at this timestep
                        # print(f'{v_n_prev = }')
                        # print(f'v_n_new = (1 - {self.α})*{v_n_prev} + {self.α}*{v_hat} = {v_n_new}\n ')

                    # print(f'S_tx_t-1 values before update:')
                    # print(f'{self.states[S_tx_prev]["value"][-5:]}')
                    # self.print_state(self.states[S_tx_prev])

        # Update value to the post-decision state directly in self.states
                    self.states[S_tx_prev]['value'] = v_n_new
                    self.states[S_tx_prev]['iteration'] += 1

                    # Update values of aggregated state
                    # a = self.agg_states[S_tx_g_prev]['count'] / agg_states_count
                    # v_g = self.agg_states[S_tx_g_prev]['value']
                    # v_0 = S_tx_prev_dict['value']
                    # v_smoothed = (1 - a) * v_g + a * v_0
                    # self.agg_states[S_tx_g_prev]['value'] = v_smoothed
                    # self.agg_states[S_tx_g_prev]['count'] += 1
                    # agg_states_count += 1

    # TRANSITION:
        # Select greedy move (if ε > 0)
                    x_ε = random.choice(X_ta)
                    x_hat = tuple((random.choices(population=[x_hat, x_ε], weights=[1 - ε, ε])[0] if self.decaying_ε
                            else random.choices(population=[x_hat, x_ε], weights=[1 - self.ε, self.ε])[0]))

        # Apply the action and get the post-decision state
                    S_tx_dict, S_tx = self.apply_action_to_state(S_t_dict, x_hat, t, n)
                    # print(f'\tE[conflicts] = {self.expected_num_conflicts(S_tx_dict)}')
                    # print(f'Initial value of S_tx: {-self.initial_value(S_tx_dict)}')
                    # print(f'Count of S_tx: {S_tx_dict['iteration']}')
                    # if self.BFA and n == self.N:
                    #     collected_features = pd.DataFrame([self.basis_features(S_tx_dict, x_hat)])
                    #     collected_features.to_csv(q, header=False, index=False)


        # Add exogeneous information to post-decision state to get the next pre-decision state
                    S_t_next_dict, S_t_next = self.add_exogeneous_info(S_tx, n)
                    # print(f'\nS_t_next:')
                    # self.print_state(S_t_next_dict)

                    self.states[S_t_next] = S_t_next_dict
                    next_state = self.states[S_t_next]

                    if self.recovered(next_state):
                        # print(next_state['t'])
                        # print(f'{self.check_disruption_(next_state['t'] + 1) = }')
                        # print(f'{self.num_conflicts(next_state) = }')
                        # self.print_state(S_tx_dict)
                        # self.print_state(next_state)

                        if n == self.N and self.plot_episode:
                            self.plot_schedule(next_state, n, self.folder)
                        break

                # Store the objective value for this iteration
                objective_value = sum(accumulated_rewards)
                objective_values[n] = objective_value
                value_function_values[n] = initial_expected_value

        self.avg_obj = np.mean(list(objective_values.values()))
        self.value_evolution = value_function_values
        self.obj_evolution = objective_values

    #################### VISUALIZE ###################
    def plot_schedule(self, state, iteration, instance):
        # Plotting
        plt.figure(figsize=(12, 5))

        # Get the states for the specified step
        t = state['t']
        current_time = self.periods[t]
        # Flags to ensure 'Unavailability' and 'Canceled Flight' are added to the legend only once
        au_label_added = False
        cancel_label_added = False

        # Plot flights based on the stored order
        for aircraft_id in self.aircraft_ids[::-1]:
            aircraft_state = state.get(aircraft_id)

            if aircraft_state:
                if aircraft_state['flights']:
                    for flight in aircraft_state['flights']:
                        ADT = flight.get('ADT')
                        AAT = flight.get('AAT')
                        flight_nr = flight.get('Flightnr')

                        if ADT and AAT:
                            # Plot the flight
                            plt.plot([ADT, AAT], [aircraft_id, aircraft_id], marker='|', color='blue',
                                     linewidth=4, markersize=10, alpha=0.5)
                            # Calculate the midpoint of the flight for labeling
                            midpoint_time = ADT + (AAT - ADT) / 2
                            # Add the flight number as a label in the middle of the flight
                            plt.text(midpoint_time, aircraft_id, flight_nr,
                                     verticalalignment='bottom', horizontalalignment='center', fontsize=10,
                                     color='black')
                else:
                    # Plot a placeholder for aircraft with no flights assigned
                    plt.plot([self.recovery_start, self.recovery_end], [aircraft_id, aircraft_id], marker='|',
                             color='gray', linewidth=2, linestyle=':')
                    plt.text(self.recovery_start, aircraft_id, 'No Flights',
                             verticalalignment='bottom', horizontalalignment='left', fontsize=8, color='gray')

        for disruption_t, aircraft in self.disruptions.items():
            potential_disruption = False  # Flag for potential disruptions
            for aircraft_id, disruptions in aircraft.items():
                if disruptions:  # Check if there are any disruptions for this aircraft
                    potential_disruption = True
                    for disruption in disruptions:
                        start_time, end_time, p, realises = disruption
                        alpha = 0.1  # Default transparency for potential disruptions
                        color = 'orange'

                        # Check disruption timing relative to the current time
                        if start_time <= current_time:
                            if realises:
                                color = 'red'  # Optional: change color for realized disruptions
                                p = 1  # Update probability to reflect realization
                            else:
                                continue

                        label_name = 'Potential Disruption' if start_time > current_time else "Realised disruption" if realises else "Realised disruption"
                        label = label_name if not au_label_added else ""
                        plt.plot([start_time, end_time], [aircraft_id, aircraft_id],
                                 color=color, linewidth=4, alpha=alpha, label=label)
                        plt.scatter([start_time, end_time], [aircraft_id, aircraft_id],
                                    color=color, marker='x', alpha=alpha, s=100)  # Markers for start and end

                        # Calculate the midpoint for labeling
                        midpoint_time = start_time + (end_time - start_time) / 100
                        plt.annotate(f'{p:.2f}',
                                     xy=(midpoint_time, aircraft_id),  # Position
                                     xytext=(0, -15),  # Offset in pixels (x=0, y=-15)
                                     textcoords='offset points',
                                     fontsize=10, color=color,
                                     ha='center', va='top',
                                     bbox=dict(facecolor='white', edgecolor='none', alpha=0.5))
                        au_label_added = True  # Only add the label once for potential disruptions

        # Plot canceled flights
        for flight, aircraft_id in self.cancelled_flights:
            ADT = flight.get('ADT')
            AAT = flight.get('AAT')
            flight_nr = flight.get('Flightnr')

            if ADT and AAT:
                # Plot the canceled flight in red with transparency
                label = 'Canceled Flight' if not cancel_label_added else ""
                plt.plot([ADT, AAT], [aircraft_id, aircraft_id], marker='|', color='darkred',
                         linewidth=4, markersize=10, alpha=0.5, label=label)
                # Calculate the midpoint of the canceled flight for labeling
                midpoint_time = ADT + (AAT - ADT) / 2
                # Add the flight number as a label in the middle of the canceled flight
                plt.text(midpoint_time, aircraft_id, flight_nr,
                         verticalalignment='bottom', horizontalalignment='center', fontsize=10,
                         color='darkred', alpha=0.9)
                cancel_label_added = True  # Only add the label once for 'Canceled Flight'

        # Retrieve the current time associated with the step
        current_time = self.periods[t] if t < len(self.periods) else self.recovery_end

        # Plot the current time as a vertical line (only once)
        plt.axvline(x=current_time, color='black', linestyle='-', linewidth=1, label='Current Time')

        # Plot recovery_start and recovery_end as vertical lines (only once)
        plt.axvline(x=self.recovery_start, color='purple', linestyle='--', linewidth=1, label='Recovery Start')
        plt.axvline(x=self.recovery_end, color='purple', linestyle='--', linewidth=1, label='Recovery End')

        plt.xlabel('Time')
        plt.ylabel('Aircraft')
        plt.title(f'Flight Schedule: t= {t}, n={iteration}, {instance}')
        plt.grid(True)

        # Format x-axis to show only time
        plt.gca().xaxis.set_major_formatter(plt.matplotlib.dates.DateFormatter('%H:%M'))
        plt.gca().xaxis.set_major_locator(mdates.HourLocator(interval=1))
        plt.xticks(rotation=45)  # Rotate x-axis labels to 45 degrees

        # Place the legend outside the plot to the right
        plt.legend(bbox_to_anchor=(1.0, 1), loc='upper left')
        plt.show()

    def print_state(self, state):
        print(f't: {state['t']}')
        for aircraft_id in self.aircraft_ids:
            print(f'\t-{aircraft_id}')
            for key, value in state[aircraft_id].items():
                print(f'\t\t-{key}: {value}')
        print(f'\t-Value = {state['value']}')
        print(f'\t-Iterations {state['iteration']}')
        state_key = self.create_hashable_state_key(state)
        # print(state_key)
        print()

    def potential_and_realised_disruptions(self):
        future_disruptions = {}
        all_disruptions = self.disruptions[self.T]

        # Iterate over each time step in self.disruptions
        for t, aircraft_disruptions in self.disruptions.items():
            future_disruptions[t] = {}

            # Iterate over each aircraft's disruptions at this time step
            for aircraft_id, disruption_list in all_disruptions.items():
                disruptions = []

                # Check each disruption
                for disruption in disruption_list:
                    start_time, end_time, prob, realised = disruption

                    # Include disruptions if they are in the future or have been realised
                    if start_time > self.periods[t]:
                        disruptions.append(disruption)

                    elif start_time <= self.periods[t]:
                        prob = 1 if realised else 0
                        disruptions.append((start_time, end_time, prob, realised))

                # Store the filtered disruptions for this aircraft
                future_disruptions[t][aircraft_id] = disruptions

        return future_disruptions

    def recovered(self, state):
        return (self.check_disruption_(state['t']) == 1 and self.num_conflicts(state) == 0)

def train_instance(instance_id):
    folder = f"TRAIN{instance_id}"
    agg_lvl = 2
    print(f"\nTraining for instance {instance_id} in folder {folder}")
    aircraft_data, flight_data, rotations_data, disruptions, recovery_start, recovery_end = read_data(folder)
    m = VFA_ADP(aircraft_data, flight_data, disruptions, recovery_start, recovery_end, agg_lvl, folder)
    TIME = time.time()
    m.train_with_vfa()
    print(f'train_with_vfa time: {time.time() - TIME}')
    return m.policy, m.states, m.agg_states, m.N, m.value_evolution, m.obj_evolution, instance_id, m.avg_obj, time.time(), m

def instance_information(instance_id):
    folder = f"TRAIN{instance_id}"
    agg_lvl = 2
    print(f"\nTraining for instance {instance_id} in folder {folder}")
    aircraft_data, flight_data, rotations_data, disruptions, recovery_start, recovery_end = read_data(folder)
    m = VFA_ADP(aircraft_data, flight_data, disruptions, recovery_start, recovery_end, agg_lvl, folder)
    stepsize = m.α if not m.harmonic_stepsize else m.harmonic_a

    return (len(m.aircraft_ids),
            len(m.flight_data),
            m.N,
            m.y,
            m.ε,
            stepsize,
            m.harmonic_stepsize,
            m.decaying_ε,
            m.pruning)

if __name__ == '__main__':
    write_results = True
    nr_instances = 1
    agg_lvl = 2
    # csv_file = '_state_features_single_p.csv'  # Define the CSV file path
    max_flights = 6

    aircraft_data, flight_data, rotations_data, disruptions, recovery_start, recovery_end = read_data('TRAIN1')
    prone_aircraft = [aircraft_data[0]['ID']]          # first aircraft is only prone to disruptions
    # initialize_csv(csv_file, [aircraft['ID'] for aircraft in aircraft_data], prone_aircraft, max_flights)

    cumulative_policy = {}
    cumulative_states = {}
    cumulative_agg_states = {}
    value_evolutions = {}
    objective_values = {}
    objective_evolutions = {}
    avg_obj = {}

    now = datetime.now()
    start_time = time.time()
    print(f'Training started at {now}')
    # Run training in parallel using ProcessPoolExecutor
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = [executor.submit(train_instance, instance_id) for instance_id in range(1, nr_instances+1)]
        # print(f'futures made in {time.time() - start_time} s')

        for index, future in enumerate(concurrent.futures.as_completed(futures)):
            TIME = time.time()
            policy, states, agg_states, N_iterations, value_evolution, obj_evolution, instance_id, obj, tim, m  = future.result()
            value_evolutions[instance_id] = value_evolution
            objective_values[instance_id] = obj
            objective_evolutions[instance_id] = obj_evolution

    # for instance_id in range(1, nr_instances + 1):
    #     policy, states, agg_states, N_iterations, value_evolution, instance_id, obj, TIME = train_instance(instance_id)
    #     value_evolutions[instance_id] = value_evolution
    #     objective_values[instance_id] = obj

    end_time = time.time()
    T = end_time - start_time
    TT = end_time - TIME
    print(f'returning values from train_instance() time: {TT} s')

    # # print('\n\n\n')
    # # save_instance(cumulative_policy, "policy")
    # # save_instance(cumulative_states, 'states')
    # # save_instance(cumulative_agg_states, 'agg_states')

    print("Training done, states and policies saved.")
    print(f"TRAINED {nr_instances} INSTANCES IN {round((T), 2)} SECONDS")
    print(f'Solved with {round((N_iterations * nr_instances) /T , 2)} iterations per second')

    F, A, N, gamma, epsilon, stepsize, harmonic, decaying, pruning = instance_information(1)

    data = (flight_data, aircraft_data, T)
    vals = (objective_values, objective_evolutions, value_evolutions)
    plot_convergence(m, nr_instances, (data), (vals))

    if write_results:
        # Define parameters and results
        params = {
            'Run': 'X',
            '|F|': F,
            '|A|': A,
            'N': N,
            'n_instances': nr_instances,
            'gamma': gamma,
            'epsilon': epsilon,
            'stepsize': stepsize,
            'harmonic': harmonic,
            'decaying_epsilon': decaying,
            'CPU': T,
            'Iterations_per_second': round((N_iterations * nr_instances) / (T), 2),
            'single_aircraft': '',
            'distr.': '',
            'action_pruning': pruning
        }

        df = pd.DataFrame([params])

        from openpyxl import load_workbook
        import os
        # Define the Excel file path
        file_path = '../Results.xlsx'
        sheet_name = 'Training'

        # Check if the file exists
        if os.path.exists(file_path):
            # If the file exists, append without overwriting
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                # Load the workbook to find the correct row
                workbook = writer.book
                if sheet_name in workbook.sheetnames:
                    # Get the last row in the existing sheet
                    start_row = workbook[sheet_name].max_row
                else:
                    # If the sheet does not exist, start from row 0
                    start_row = 0
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=(start_row == 0), startrow=start_row)
        else:
            # If the file doesn't exist, create it and write the dataframe
            df.to_excel(file_path, sheet_name=sheet_name, index=False)

        print("Results and parameters saved to Excel.")


  # Snapshot of memory usage
  #   snapshot = tracemalloc.take_snapshot()
  #   top_stats = snapshot.statistics('lineno')

    # print("[ Top 10 Memory Consuming Lines ]")
    # for stat in top_stats[:10]:
    #     print(stat)